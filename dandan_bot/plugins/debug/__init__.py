import nonebot
from nonebot.adapters.onebot.v11 import *
from nonebot import on_command
from nonebot.rule import to_me
import os
import openpyxl
from pathlib import Path

coin_per_capso = on_command("debug->coin_per_capso", priority=2, rule=to_me())
__user: dict[int] = nonebot.get_driver().config.debugger_user


@coin_per_capso.handle()
async def check_and_send(bot: Bot, event: Event):
    __id = event.get_user_id()
    for debugger_id in __user:
        if __id == debugger_id:
            await coin_per_capso.send('您好，debugger!')
            str_path = Path(os.getcwd())
            sx = openpyxl.load_workbook(str_path / 'dandan_bot' / 'libraries' / 'user.xlsx')
            sheet = sx['Sheet1']
            for x in range(2, sheet.max_row + 1):
                if sheet.cell(row=x, column=1).value == str(__id):
                    sheet.cell(row=x, column=2).value = str(int(sheet.cell(row=x, column=2).value) + 15)
                    sx.save(str_path / 'dandan_bot' / 'libraries' / 'user.xlsx')
                    await coin_per_capso.finish('您的账号现已增加15璇璇币，请您收好并谨慎使用~')
            await coin_per_capso.finish('作为一个Debugger居然没有账号？！难以相信……')
    await coin_per_capso.finish('这是什么啊？是什么暗号吗？我不懂欸~')