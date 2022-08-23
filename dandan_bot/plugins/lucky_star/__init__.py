# coding=gbk
import asyncio
import os
import random
from pathlib import Path
from typing import Type

import openpyxl
from nonebot import on_command
from nonebot.adapters import Message
from nonebot.adapters.onebot.v11 import *
from nonebot.params import Depends
from nonebot.rule import to_me

capso = on_command("�齱", rule=to_me(), priority=5, aliases={'capso', 'lucky_star'})


class EventChecker:
    def __init__(self, EventClass: Type[MessageEvent]):
        self.event_class = EventClass

    def __call__(self, event: MessageEvent) -> bool:
        return isinstance(event, self.event_class)


checker = EventChecker(GroupMessageEvent)


@capso.handle()
async def first_handle_receive(bot: Bot, event: Event, flag: bool = Depends(checker)):
    str_path = Path(os.getcwd())
    sx = openpyxl.load_workbook(str_path / 'dandan_bot' / 'libraries' / 'user.xlsx')
    sheet = sx['Sheet1']
    for x in range(1, sheet.max_row + 1):
        if sheet.cell(row=x, column=1).value == str(event.get_user_id()):
            if int(sheet.cell(row=x, column=2).value) >= 15:
                sheet.cell(row=x, column=2).value = str(int(sheet.cell(row=x, column=2).value) - 15)
                __id = sheet.cell(row=x, column=1).value
                get = ''
                adder = 0
                fflag = False
                nz_flag = False
                dd_flag = False
                mmmmmmsg = ''
                tmp = random.randint(1, 10000000)
                if tmp <= 1500:
                    get = '������Ա����!!!'
                elif tmp <= 200001:
                    get = '60�诱�,���Ǳ���������һ��!'
                    adder = 30
                elif tmp <= 700001:
                    get = '��Ⱥ��Ůװ�Ĵ����ȣ�!'
                    nz_flag = True
                elif tmp <= 1200001:
                    get = '�����Ա�Ʋ����Ļ��ᣡ����ϲ'
                elif tmp <= 1700001:
                    get = '������Ц�ݣ�\n����20�诱��������£�'
                    adder = 20
                elif tmp <= 2200001:
                    get = 'EK³�ȵ�Ůװһ�ݣ�\n����20�诱��������£�'
                    adder = 20
                elif tmp <= 3000001:
                    get = '�Х����ղصĵ�������һ�ݣ�����ͫ��ϲ\n'
                    dd_flag = True
                elif tmp <= 4000001:
                    get = 'ȥ���������ξ�,���ǹ�����!'
                elif tmp <= 5000001:
                    get = '�ж��߱���\n�����ĸ���棡'
                elif tmp <= 6000001:
                    get = '�������ƶ�������������ϲ'
                elif tmp <= 6500001:
                    get = '114514��...'
                    fflag = True
                elif tmp <= 8000001:
                    get = '����!'
                    tmp = random.randint(1, 10000000)
                    if tmp <= 5:
                        adder = 9999
                    elif tmp <= 100:
                        adder = 616
                    elif tmp <= 9000000:
                        adder = -random.randint(5, 10)
                    else:
                        adder = 10 * random.randint(1, 10)
                    if adder > 0:
                        mmmmmmsg = '�����³���' + str(adder) + '���诱�!'
                    elif adder < 0:
                        mmmmmmsg = '���γԵ���' + str(-adder) + '���诱�!'
                    else:
                        mmmmmmsg = '������æ����Ǯ��һ���诱Ҷ��������~'
                else:
                    adder = random.randint(10, 25)
                    get = str(adder) + '���诱�!'

                sheet.cell(row=x, column=2).value = str(int(sheet.cell(row=x, column=2).value) + adder)
                if int(sheet.cell(row=x, column=2).value) < 0:
                    sheet.cell(row=x, column=2).value = '0'
                sx.save(str_path / 'dandan_bot' / 'libraries' / 'user.xlsx')
                msg: Message = MessageSegment.at(user_id=__id) + '\n'
                msg1 = '��鵽��' + get
                if flag:
                    print(1)
                    await capso.send(msg + msg1)
                else:
                    await capso.send(msg1)
                if fflag:
                    await asyncio.sleep(2)
                    await capso.finish('�ɱ���')
                if dd_flag:
                    image_path: Path = str_path / 'data' / 'images' / 'dandan.png'
                    msg2 = MessageSegment.image(file=image_path)
                    await capso.finish(msg2)
                if nz_flag:
                    await asyncio.sleep(0.5)
                    if flag:
                        __info = await bot.get_group_member_list(group_id=event.group_id)
                        for __user in __info:
                            if (await bot.get_group_member_info(group_id=event.group_id,user_id=__user["user_id"],no_cache=True))["role"] == 'owner':
                                owner_id = __user["user_id"]
                                await capso.finish(MessageSegment.at(int(owner_id)) + '���Ůװ~')
                    else:
                        await capso.finish("����������˽�ĚG~���˰ɩ�(���� ��)��")
                if mmmmmmsg != '':
                    await asyncio.sleep(0.1)
                    await capso.finish(mmmmmmsg)
                await capso.finish('�ڴ����´�����~���Ϲ���')
            else:
                await capso.finish('�ٳ齱��ҪǷǮ��~')

    await capso.finish('����˭�����Һ���̫��ʶ��������ء�������ע��������~')
