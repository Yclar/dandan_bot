from nonebot.adapters import Bot, Event, Message
from nonebot.adapters.onebot.v11 import MessageSegment, GroupMessageEvent, MessageEvent
from nonebot import on_command
from nonebot.rule import to_me
from nonebot.typing import T_State
from nonebot.params import State, CommandArg
import os
import openpyxl
from pathlib import Path
from nonebot.params import Depends
from typing import Type

personal_info = on_command("个人信息", rule=to_me(), priority=5, aliases={'personal_info', 'p_info', 'pinfo'})


class EventChecker:
    def __init__(self, EventClass: Type[MessageEvent]):
        self.event_class = EventClass

    def __call__(self, event: MessageEvent) -> bool:
        return isinstance(event, self.event_class)


checker = EventChecker(GroupMessageEvent)


@personal_info.handle()
async def first_handle_receive(bot: Bot, event: Event, flag: bool = Depends(checker)):

    str_path = Path(os.getcwd())
    sx = openpyxl.load_workbook(str_path / 'dandan_bot' / 'libraries' / 'user.xlsx')
    sheet = sx['Sheet1']
    for x in range(2, sheet.max_row + 1):
        if sheet.cell(row=x, column=1).value == str(event.get_user_id()):
            __id = sheet.cell(row=x, column=1).value
            __coin = sheet.cell(row=x, column=2).value
            __sign_time = sheet.cell(row=x, column=3).value
            msg: Message = MessageSegment.at(user_id=__id) + '\n'
            msg1 = '你的QQ号是' + str(__id) + '\n'
            msg2 = '你现在一共有' + str(__coin) + '个璇璇币\n'
            msg3 = '你上一次签到是在' + str(__sign_time)
            if flag:
                print(1)
                await personal_info.finish(msg + msg1 + msg2 + msg3)
            else:
                await personal_info.finish(msg1 + msg2 + msg3)
    await personal_info.finish('你是谁啊？我好像不太认识你的样子呢……请先注册再来吧~')
