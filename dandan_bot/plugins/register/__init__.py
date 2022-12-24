from asyncio import sleep
from nonebot.adapters import Bot, Event
from nonebot import on_command, on_notice
from nonebot.rule import to_me
from os import getcwd
from openpyxl import load_workbook
from pathlib import Path


register = on_command("注册", rule=to_me(), priority=5, block=True)
group_register = on_notice(priority=1)


@register.handle()
async def first_handle_receive(event: Event):
    str_path = Path(getcwd())
    sx = load_workbook(str_path / 'dandan_bot' / 'libraries' /'user.xlsx')
    sheet = sx['Sheet1']
    for x in range(2, sheet.max_row + 1):
        if sheet.cell(row=x, column=1).value == str(event.get_user_id()):
            await register.send("小笨蛋~")
            await sleep(0.5)
            await register.finish("你已经注册过了~")
    for x in range(2, sheet.max_row + 2):
        if sheet.cell(row=x, column=1).value is None:
            sheet.cell(row=x, column=1).value = str(event.get_user_id())
            sheet.cell(row=x, column=2).value = str(0)
            sx.save(str_path / 'dandan_bot' / 'libraries' / 'user.xlsx')
            await register.finish("注册成功~")


@group_register.handle()
async def second_handle_receive(event: Event, bot: Bot):
    if event.notice_type == "group_increase" and event.user_id == int(bot.self_id):
        str_path = Path(getcwd())
        sx = load_workbook(str_path / 'dandan_bot' / 'libraries' / 'group.xlsx')
        sheet = sx['Sheet1']
        for x in range(2, sheet.max_row + 1):
            if sheet.cell(row=x, column=1).value == str(event.group_id):
                await group_register.finish()
        for x in range(2, sheet.max_row + 2):
            if sheet.cell(row=x, column=1).value is None:
                sheet.cell(row=x, column=1).value = str(event.group_id)
                sheet.cell(row=x, column=2).value = str("欢迎新成员~")
                sx.save(str_path / 'dandan_bot' / 'libraries' / 'group.xlsx')
                await group_register.finish()