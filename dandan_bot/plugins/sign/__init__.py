from nonebot.adapters import Event
from nonebot import on_command
from nonebot.rule import to_me
from os import getcwd
from openpyxl import load_workbook
from pathlib import Path
from random import randint
from datetime import datetime


sign = on_command("签到", rule=to_me(), priority=5, block=True, aliases={"早"})

@sign.handle()
async def first_handle_receive(event: Event):
    str_path = Path(getcwd())
    sx = load_workbook(str_path / 'dandan_bot' / 'libraries' / 'user.xlsx')
    sheet = sx['Sheet1']
    for x in range(2, sheet.max_row + 1):
        if sheet.cell(row=x, column=1).value == str(event.get_user_id()):
            adder: int = randint(10, 25)
            dt = datetime.today()
            if str(sheet.cell(row=x, column=3).value) == str(dt.date()):
                sx.save(str_path / 'dandan_bot' / 'libraries' / 'user.xlsx')
                await sign.finish('欸？怎么又是你？别来rua羊毛了好不好uwu~')
            sheet.cell(row=x, column=2).value = str(int(sheet.cell(row=x, column=2).value) + adder)
            sheet.cell(row=x, column=3).value = str(dt.date())
            sx.save(str_path / 'dandan_bot' / 'libraries' / 'user.xlsx')
            await sign.finish("恭喜你获得了" + str(adder) + "个璇璇币，请明天再来哦~")
    sx.save(str_path / 'dandan_bot' / 'libraries' / 'user.xlsx')
    await sign.finish("欸？Σ(っ °Д °;)っ这名单上没有你的名字！注册完再来吧~")