# coding=gbk
from os import getcwd
from openpyxl import load_workbook
from pathlib import Path
from typing import Optional


async def get(song_name: str) -> Optional[str]:
    str_path = Path(getcwd())
    sx = load_workbook(str_path / 'dandan_bot' / 'libraries' / 'song_diff.xlsx')
    phi = sx['phi']
    ret = []
    cnt = -1
    for x in range(2, phi.max_row + 1):
        f = 0
        alias = str(phi.cell(row=x, column=9).value).split()
        for i in alias:
            if str(i).title() == song_name.title():
                f = 1
        if (str(phi.cell(row=x, column=5).value).title().strip() == song_name.title().strip()) or f:
            cnt = cnt + 1
            __lv = str(phi.cell(row=x, column=6).value)
            __diff = str(phi.cell(row=x, column=7).value)
            __n = str(phi.cell(row=x, column=8).value)
            __ver = str(phi.cell(row=x, column=4).value)
            __name = str(phi.cell(row=x, column=5).value)
            ret.append('')
            ret[cnt] += __lv + '#' + __diff + '#' + __n + '#' + __ver + '#' + __name + '#@#'
            print(ret)
    return 'baka'


async def alias_(alias: str, name: str) -> Optional[str]:
    str_path = Path(getcwd())
    sx = load_workbook(str_path / 'dandan_bot' / 'libraries' / 'song_diff.xlsx')
    phi = sx['phi']
    ret = 'baka'
    for x in range(2, phi.max_row + 1):
        if str(phi.cell(row=x, column=5).value) == name:
            ret = 'ok'
            if phi.cell(row=x, column=9).value == None:
                phi.cell(row=x, column=9).value = alias + ' '
            else:
                phi.cell(row=x, column=9).value = str(phi.cell(row=x, column=9).value) + alias + ' '
    sx.save(str_path / 'dandan_bot' / 'libraries' / 'song_diff.xlsx')
    return ret
