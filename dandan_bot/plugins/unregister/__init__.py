from asyncio import sleep
from nonebot.adapters import Event, Message
from nonebot import on_command
from nonebot.rule import to_me
from nonebot.params import Arg
from os import getcwd
from openpyxl import load_workbook
from pathlib import Path


unregister = on_command("注销", rule=to_me(), priority=5, block=True)


@unregister.handle()
async def first_handle_receive():
    await unregister.send('欸？！你真的要这样做吗？如果继续的话请输入"是"，否则就输入别的~')


@unregister.got("ans")
async def final_operation(event: Event, ans: Message = Arg()):
    if str(ans) == '是':
        str_path = Path(getcwd())
        sx = load_workbook(str_path / 'dandan_bot' / 'libraries' / 'user.xlsx')
        sheet = sx['Sheet1']
        for x in range(1, sheet.max_row + 1):
            if sheet.cell(row=x, column=1).value == str(event.get_user_id()):
                for y in range(1, sheet.max_column + 1):
                    sheet.cell(row=x, column=y).value = None
                sx.save(str_path / 'dandan_bot' / 'libraries' / 'user.xlsx')
                await unregister.finish("名单上没有你的名字了哦（哭）")
        await unregister.send("小笨蛋~")
        await sleep(0.5)
        await unregister.finish("你还没有注册~")
    else:
        await unregister.finish("阿拉，你逗我呢，吓死我了＞︿＜")
