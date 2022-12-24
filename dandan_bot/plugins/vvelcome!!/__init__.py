from nonebot.adapters import Event, Bot
from nonebot import on_notice, on_command
from nonebot.adapters.onebot.v11 import MessageSegment, Message
from os import getcwd, sep
from openpyxl import load_workbook
from pathlib import Path
from nonebot.params import CommandArg, State
from nonebot.rule import to_me
from nonebot.typing import T_State


vvelcome = on_notice(priority=2)
set_welcome = on_command("设置欢迎词", rule=to_me(), priority=5, block=True)
vision_welcome = on_command("显示欢迎词", rule=to_me(), priority=5, block=True)


@vvelcome.handle()
async def send_vvelcome(bot: Bot, event: Event):
    print(1)
    print(event)
    if event.notice_type == "group_increase" and event.user_id != int(bot.self_id):
        welcomeMessage = ''
        str_path = Path(getcwd())
        sx = load_workbook(str_path / 'dandan_bot' / 'libraries' / 'group.xlsx')
        sheet = sx['Sheet1']
        for x in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=x, column=1).value) == str(event.group_id):
                welcomeMessage = sheet.cell(row=x, column=2).value
        print('a', str(event.group_id), 'a', sep='')
        print(welcomeMessage)
        await vvelcome.finish(MessageSegment.at(int(event.user_id)) + "\n" + str(welcomeMessage))


@set_welcome.handle()
async def first_handle_receive(bot: Bot, event: Event, state: T_State = State(), args: Message = CommandArg()):
    user_role = (await bot.get_group_member_info(group_id=event.group_id, user_id=event.user_id, no_cache=True))["role"]
    if user_role == 'owner' or user_role == 'admin':
        plain_text = args.extract_plain_text()
        if plain_text:
            state["_data_"] = plain_text
    else:
        await set_welcome.finish("您不是群主或者管理员哦~")


@set_welcome.got("_data_", prompt="请问有什么内容？")
async def data_input(event: Event, state: T_State = State()):
    str_path = Path(getcwd())
    sx = load_workbook(str_path / 'dandan_bot' / 'libraries' / 'group.xlsx')
    sheet = sx['Sheet1']
    for x in range(2, sheet.max_row + 1):
        if sheet.cell(row=x, column=1).value == str(event.group_id):
            sheet.cell(row=x, column=2).value = state["_data_"]
            break
    sx.save(str_path / 'dandan_bot' / 'libraries' / 'group.xlsx')
    await set_welcome.finish("欢迎词设置成功！")


@vision_welcome.handle()
async def first_handle_receive_another(bot: Bot, event: Event):
    user_role = (await bot.get_group_member_info(group_id=event.group_id, user_id=event.user_id, no_cache=True))["role"]
    if user_role == 'owner' or user_role == 'admin':
        str_path = Path(getcwd())
        sx = load_workbook(str_path / 'dandan_bot' / 'libraries' / 'group.xlsx')
        sheet = sx['Sheet1']
        for x in range(2, sheet.max_row + 1):
            if sheet.cell(row=x, column=1).value == str(event.group_id):
                __data = sheet.cell(row=x, column=2).value
                await vision_welcome.finish('目前欢迎词为：\n' + str(__data))
    else:
        await set_welcome.finish("您不是群主或者管理员哦~")