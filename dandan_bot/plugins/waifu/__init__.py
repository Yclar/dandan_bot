# coding=gbk
from nonebot.adapters.onebot.v11 import *
from nonebot import on_command
from nonebot.rule import to_me
from nonebot.params import Depends
from typing import Type
import random
import os
import openpyxl
from pathlib import Path
from datetime import datetime

from pyparsing import col


class EventChecker:
    def __init__(self, EventClass: Type[MessageEvent]):
        self.event_class = EventClass

    def __call__(self, event: MessageEvent) -> bool:
        return isinstance(event, self.event_class)


checker = EventChecker(GroupMessageEvent)

waifu = on_command('贴贴', aliases={'抽老婆', 'waifu'}, priority=5, rule=to_me())


@waifu.handle()
async def check_and_send(bot: Bot, event: Event, flag: bool = Depends(checker)):
    if flag:
        str_path = Path(os.getcwd())
        sx = openpyxl.load_workbook(str_path / 'dandan_bot' / 'libraries' / 'user.xlsx')
        sheet_info = sx['Sheet1']
        dt = str(datetime.today().date())
        id = -1
        for x in range(1, sheet_info.max_row + 1):
            if sheet_info.cell(row=x, column=1).value == str(event.get_user_id()):
                id = x
                break
        if id == -1:
            sx.save(str_path / 'dandan_bot' / 'libraries' / 'user.xlsx')
            await waifu.finish("欸？Σ(っ °Д °;)っ这名单上没有你的名字！注册完再来吧~")

        all_waifu = await bot.get_group_member_list(group_id=event.group_id, type='all')
        your_waifu = random.choice(all_waifu)
        while your_waifu['user_id'] == event.user_id or your_waifu['user_id'] == event.self_id:
            your_waifu = random.choice(all_waifu)

        sheet_waifu = sx['Sheet2']
        f = 1
        for x in range(1, sheet_waifu.max_column + 1):
            waifu_info = sheet_waifu.cell(row=id, column=x).value
            if waifu_info is None:
                continue
            pos1 = waifu_info.find('@')
            pos2 = waifu_info.find("#")
            waifu_group = int(waifu_info[0: pos1])
            waifu_time = waifu_info[pos2 + 1:]
            waifu_userid = int(waifu_info[pos1 + 1: pos2])
            if waifu_group == event.group_id:
                f = 0
                if waifu_time == dt:
                    your_waifu = await bot.get_group_member_info(
                        group_id=event.group_id, user_id=waifu_userid, no_cache=True
                    )
                else:
                    sheet_waifu.cell(row=id, column=x).value = \
                        str(event.group_id) + '@' + str(your_waifu['user_id']) + '#' + dt
        if f:
            for x in range(1, sheet_waifu.max_column + 1):
                if sheet_waifu.cell(row=id, column=x).value is None:
                    sheet_waifu.cell(row=id, column=x).value = \
                        str(event.group_id) + '@' + str(your_waifu['user_id']) + '#' + dt
                    break

        nickname = str(your_waifu['nickname'])
        userid = str(your_waifu['user_id'])
        avatar_url = 'http://q1.qlogo.cn/g?b=qq&nk=' + userid + '&s=3'
        msg0 = MessageSegment.at(event.user_id)
        msg1 = '\n今天你亲爱的群友是\n'
        msg2 = MessageSegment.image(file=avatar_url)
        msg3 = '【' + nickname + '】(' + userid + ')哒！'
        sx.save(str_path / 'dandan_bot' / 'libraries' / 'user.xlsx')
        await waifu.finish(msg0 + msg1 + msg2 + msg3)

    else:
        await waifu.finish("但是我们是私聊欸~算了吧┑(￣Д ￣)┍")